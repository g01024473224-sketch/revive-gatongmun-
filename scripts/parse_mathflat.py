"""
매쓰플랫 Excel 파서
- 입력: 이상준선생님_{학년}_{YYYY년 MM월}_학습지 학습내역.xlsx
- 출력: 학생별 주간평가/선행평가 JSON

주요 규칙:
- 주간 TEST / 주간평가 / 월간 TEST → 주간평가 테이블
- 단원 TEST → 선행평가 테이블
- 나머지(연습문제, 기본, 교재오답, 학습지오답, 연산, 숙제, 유형별학습 등) → 제외
- 재시 차수는 학습지명의 "재시/재재시/재재재시..." 로 판정 (괄호 유무 무관)
- 점수가 '-' 또는 None 또는 채점 문항 수 0 → 무효 레코드
"""

import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ---------- 태그 → 가통문 카테고리 매핑 ----------
WEEKLY_EVAL_TAGS = {"주간 TEST", "주간평가", "월간 TEST", "월간TEST", "주간TEST"}
UNIT_EVAL_TAGS = {"단원 TEST", "단원TEST", "단원평가"}


# ---------- 유틸 ----------
def count_retry(name: str) -> int:
    """학습지명에서 '재시' 앞의 '재' 개수를 셈 → 재시 차수 반환 (0=원본, 1=재시, 2=재재시...)"""
    # "재재재시", "(재시)", " 재시" 등 모두 매칭
    m = re.search(r"(재+)시", name)
    if not m:
        return 0
    return len(m.group(1))  # "재" 문자 개수 = 재시 차수


def strip_retry_suffix(name: str) -> str:
    """학습지명에서 재시 표시 제거 → 원본 단원명 복원."""
    # "(재시)", "(재재시)", " 재시", " 재재시" 등 제거
    cleaned = re.sub(r"\s*[\(（]?\s*재+시\s*[\)）]?\s*$", "", name).strip()
    return cleaned


def parse_week_from_name(name: str) -> int | None:
    """학습지명에서 'N월 M주' 패턴 추출 → 주차 번호(M) 반환."""
    m = re.search(r"(\d+)월\s*(\d+)주", name)
    if m:
        return int(m.group(2))
    return None


def parse_week_from_date(d) -> int | None:
    """날짜에서 해당 월의 주차 계산 (1일=1주차, 8일=2주차...)"""
    if d is None:
        return None
    if isinstance(d, str):
        try:
            d = datetime.fromisoformat(d.replace("/", "-").strip()).date()
        except ValueError:
            return None
    if isinstance(d, datetime):
        d = d.date()
    if not isinstance(d, date):
        return None
    return (d.day - 1) // 7 + 1


def normalize_score(val) -> int | None:
    """점수값 정규화 ('-', None, 빈칸 → None)."""
    if val is None or val == "" or val == "-":
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


# ---------- 메인 파서 ----------
def parse_mathflat_excel(filepath: str | Path) -> dict:
    """
    매쓰플랫 Excel 하나를 파싱해서 학생별 평가 데이터 반환.

    반환 형식:
    {
      "source_file": "...",
      "period": "2026년 03월",
      "school_level": "중학교",
      "grade": "1",
      "students": {
        "고하율": {
          "weekly_evals": [
            {"단원명": "...", "주차": 2, "재시차수": 0, "점수": 85,
             "날짜": "2026-03-11", "원본_학습지명": "..."},
            ...
          ],
          "unit_evals": [
            {"단원명": "...", "재시차수": 0, "점수": 70, "날짜": "...", ...}
          ]
        }
      },
      "excluded_tags_count": {"연습문제": 52, "기본": 25, ...}
    }
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 헤더에서 컬럼 위치 탐색
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and "학생 이름" in (row or ()):
            header_row = idx
            header = list(row)
            break
    if header_row is None:
        raise ValueError(f"헤더를 찾을 수 없음: {filepath}")

    col = {name: i for i, name in enumerate(header) if name}

    required = ["기간", "학교급", "학년", "학생 이름", "학습지 태그",
                "학습지 명", "학습지 출제일", "채점 문항 수", "점수"]
    for r in required:
        if r not in col:
            raise ValueError(f"컬럼 누락: {r} in {filepath}")

    result = {
        "source_file": filepath.name,
        "period": None,
        "school_level": None,
        "grade": None,
        "students": defaultdict(lambda: {
            "weekly_evals": [],
            "unit_evals": [],
            "all_activities": [],  # 난이도 분석용: 모든 학습지 기록
        }),
        "excluded_tags_count": defaultdict(int),
    }

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[col["학생 이름"]]:
            continue

        tag = (row[col["학습지 태그"]] or "").strip()
        student = str(row[col["학생 이름"]]).strip()
        raw_name = str(row[col["학습지 명"]] or "").strip()
        date_val = row[col["학습지 출제일"]]
        scored = row[col["채점 문항 수"]]
        score = normalize_score(row[col["점수"]])
        난이도 = row[col["난이도"]]
        전체_문항 = row[col["전체 문항 수"]]

        # 메타정보 저장
        if result["period"] is None:
            result["period"] = row[col["기간"]]
            result["school_level"] = row[col["학교급"]]
            result["grade"] = str(row[col["학년"]])

        # 채점되지 않은 항목(점수 무효) 제외
        if score is None or (isinstance(scored, (int, float)) and scored == 0):
            continue

        retry_n = count_retry(raw_name)
        clean_name = strip_retry_suffix(raw_name)
        week = parse_week_from_name(raw_name) or parse_week_from_date(date_val)

        if isinstance(date_val, datetime):
            date_str = date_val.date().isoformat()
        elif isinstance(date_val, date):
            date_str = date_val.isoformat()
        else:
            date_str = str(date_val) if date_val else None

        # 난이도 정규화 (1~5 숫자로)
        try:
            난이도_int = int(난이도) if 난이도 is not None else None
        except (ValueError, TypeError):
            난이도_int = None

        try:
            문항수 = int(전체_문항) if 전체_문항 is not None else 0
        except (ValueError, TypeError):
            문항수 = 0

        eval_record = {
            "단원명": clean_name,
            "재시차수": retry_n,
            "점수": score,
            "날짜": date_str,
            "원본_학습지명": raw_name,
            "난이도": 난이도_int,
            "태그": tag,
            "문항수": 문항수,
        }
        if week is not None:
            eval_record["주차"] = week

        # 모든 활동 기록 (난이도 분석용)
        result["students"][student]["all_activities"].append(eval_record)

        # 가통문 표시용 분류
        if tag in WEEKLY_EVAL_TAGS:
            result["students"][student]["weekly_evals"].append(eval_record)
        elif tag in UNIT_EVAL_TAGS:
            result["students"][student]["unit_evals"].append(eval_record)
        else:
            result["excluded_tags_count"][tag] += 1

    # defaultdict → dict 변환
    result["students"] = {k: v for k, v in result["students"].items()}
    result["excluded_tags_count"] = dict(result["excluded_tags_count"])
    return result


# ---------- CLI ----------
def main():
    if len(sys.argv) < 2:
        print("사용법: python parse_mathflat.py <엑셀파일경로> [<출력JSON경로>]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = (
        Path(sys.argv[2]) if len(sys.argv) >= 3
        else input_path.with_suffix(".parsed.json")
    )

    data = parse_mathflat_excel(input_path)

    # UTF-8 JSON 출력 (한글 보존)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    # 요약 출력
    print(f"✅ 파싱 완료: {input_path.name}")
    print(f"   기간: {data['period']}  학교급: {data['school_level']}  학년: {data['grade']}")
    print(f"   학생 수: {len(data['students'])}")
    total_weekly = sum(len(s['weekly_evals']) for s in data['students'].values())
    total_unit = sum(len(s['unit_evals']) for s in data['students'].values())
    print(f"   주간평가: {total_weekly}건  |  선행평가: {total_unit}건")
    print(f"   제외된 태그: {data['excluded_tags_count']}")
    print(f"   출력: {output_path}")


if __name__ == "__main__":
    main()
