"""
가통문 Excel 렌더러 (레퍼런스 레이아웃 + A4 인쇄 최적화)

- 총 13개 컬럼(A-M) 고정 레이아웃
  · A: 주차/학년 라벨 (6.5)
  · B-E: 수업계획 본문 (13 × 4)  — 본행/선행 2쌍 (B-C, D-E)
  · F: 세로 여백 (1.75)
  · G-M: 달력 7요일 (8.5 × 7)
- A4 세로, 1페이지 맞춤 인쇄

레이아웃 A (평가 있음):
  상단: 좌측 수업계획 (A-E) + 우측 작은 달력 (G-M)
  하단: 주간평가 / 선행평가 표

레이아웃 B (평가 없음):
  상단: 수업계획 전체 폭으로 확장 (A-M)
  하단: 큰 달력 (A-M)
"""

from __future__ import annotations

import calendar
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------- 픽셀 ↔ Excel 단위 변환 (원장님 Excel 실측) ----------
# 실측 (2026-04-17):
#   width 8.17 → 98px, width 5.25 → 63px
#   → 공식: pixels = width × 12, width = pixels / 12
def px_to_col_width(px: float) -> float:
    return round(px / 12, 2)

def px_to_row_height(px: float) -> float:
    # 실측 (2026-04-17): 19pt→38px, 25.5pt→51px, 45pt→90px
    # 완벽히 1pt = 2px → pt = px / 2
    return round(px / 2, 2)

# ---------- 컬럼 폭 (원장님 지정: B~E=105px, A/F~M=70px) ----------
_W_105 = px_to_col_width(105)  # 8.75 (원장님 Excel 기준)
_W_70 = px_to_col_width(70)    # 5.83

COL_WIDTHS = {
    "A": _W_70,
    "B": _W_105, "C": _W_105, "D": _W_105, "E": _W_105,
    "F": _W_70,
    "G": _W_70, "H": _W_70, "I": _W_70, "J": _W_70,
    "K": _W_70, "L": _W_70, "M": _W_70,
}

# ---------- 행 높이 (원장님 실측 기준, 2026-04-17) ----------
ROW_H_STANDARD = px_to_row_height(35)      # 17.5pt — 일반 내용 (35px)
ROW_H_NOTICE_BODY = px_to_row_height(105)  # 52.5pt — 공지 본문 (105px)
ROW_H_COMMENT_BODY = px_to_row_height(230) # 115.0pt — 코멘트 본문 (230px)
ROW_H_TITLE = px_to_row_height(39)         # 19.5pt — 타이틀 각 행 (39px)
# 공백 행(15pt = 30px)은 Excel 기본값이라 별도 설정 불필요

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

NAVY_FILL = PatternFill("solid", fgColor="1F3864")
LIGHT_FILL = PatternFill("solid", fgColor="F2F2F2")
WARNING_FILL = PatternFill("solid", fgColor="FFF4CC")

# 폰트 크기: 제목 10pt, 내용 9pt, 작은 달력 8pt (원장님 지정)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=48)
NAME_LABEL_FONT = Font(name="맑은 고딕", bold=True, size=10)
NAME_VALUE_FONT = Font(name="맑은 고딕", size=9)
SECTION_TITLE_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
ROLE_HEADER_FONT = Font(name="맑은 고딕", bold=True, size=10)
TEXTBOOK_FONT = Font(name="맑은 고딕", bold=True, color="CC0066", size=9)
WEEK_LABEL_FONT = Font(name="맑은 고딕", size=9)
CELL_FONT = Font(name="맑은 고딕", size=9)
RED_FONT = Font(name="맑은 고딕", color="CC0000", bold=True, size=9)
CAL_DAY_FONT = Font(name="맑은 고딕", size=8)
CAL_SUN_FONT = Font(name="맑은 고딕", size=8, color="CC0000")
CAL_SAT_FONT = Font(name="맑은 고딕", size=8, color="0066CC")
CAL_HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
TOPLEFT = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)


# ---------- 유틸 ----------
def setup_page(ws: Worksheet) -> None:
    """A4 세로, 1페이지 맞춤, 좁은 여백."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.horizontalCentered = True
    # 인쇄 영역: A-M
    ws.print_area = "A1:M60"

    # 컬럼 폭 설정
    for letter, w in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w


def merge(ws: Worksheet, rng: str, value="", font=None, fill=None, align=CENTER, border=True):
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    cell = ws[top]
    cell.value = value
    cell.alignment = align
    if font: cell.font = font
    if fill: cell.fill = fill
    if border:
        for row in ws[rng]:
            for c in row:
                c.border = BORDER_THIN


def set_cell(ws: Worksheet, addr: str, value="", font=None, fill=None, align=CENTER, border=True):
    cell = ws[addr]
    cell.value = value
    cell.alignment = align
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = BORDER_THIN


# ---------- 섹션: 타이틀/헤더 ----------
def render_title(ws: Worksheet) -> int:
    """Revive Math 대제목 (48pt). 끝난 다음 row 반환."""
    merge(ws, "A1:M3", "Revive Math", font=TITLE_FONT,
          align=Alignment(horizontal="left", vertical="center", indent=1),
          border=False)
    ws.row_dimensions[1].height = ROW_H_TITLE
    ws.row_dimensions[2].height = ROW_H_TITLE
    ws.row_dimensions[3].height = ROW_H_TITLE
    # 하단 굵은 밑줄
    for col in "ABCDEFGHIJKLM":
        cell = ws[f"{col}3"]
        cell.border = Border(bottom=Side(style="medium", color="000000"))
    return 5  # 빈 줄 후


def render_name_date(ws: Worksheet, start_row: int, student: dict, year: int, month: int) -> int:
    """NAME / DATE / ☺ 한 줄. 끝난 다음 row."""
    r = start_row
    ws.row_dimensions[r].height = ROW_H_STANDARD

    set_cell(ws, f"A{r}", "NAME", font=NAME_LABEL_FONT, fill=LIGHT_FILL)
    merge(ws, f"B{r}:E{r}", f"{student['code']}_{student['name']}",
          font=NAME_VALUE_FONT)
    set_cell(ws, f"F{r}", "", border=False)
    set_cell(ws, f"G{r}", "DATE:", font=NAME_LABEL_FONT, fill=LIGHT_FILL)
    merge(ws, f"H{r}:L{r}", f"{year}.{month:02d}", font=NAME_VALUE_FONT)
    set_cell(ws, f"M{r}", "☺", font=Font(name="맑은 고딕", size=14))

    return r + 2  # 다음 섹션 전 간격


def render_warnings(ws: Worksheet, start_row: int, warnings: list[str]) -> int:
    if not warnings:
        return start_row
    r = start_row
    ws.row_dimensions[r].height = 20
    merge(ws, f"A{r}:M{r}", "⚠️ " + " | ".join(warnings),
          font=Font(name="맑은 고딕", bold=True, color="AA4400", size=10),
          fill=WARNING_FILL,
          align=Alignment(horizontal="left", vertical="center", indent=1))
    return r + 1


def render_notice(ws: Worksheet, start_row: int, notice_text: str) -> int:
    """공지사항 섹션 (학원 휴일, 방학, 시험 일정 등)."""
    r = start_row
    # 섹션 제목
    ws.row_dimensions[r].height = ROW_H_STANDARD
    merge(ws, f"A{r}:M{r}", "공지사항",
          font=SECTION_TITLE_FONT, fill=NAVY_FILL)
    r += 1
    # 본문: 105px
    ws.row_dimensions[r].height = ROW_H_NOTICE_BODY
    body = notice_text.strip() if notice_text else "● "
    merge(ws, f"A{r}:M{r}", body,
          font=Font(name="맑은 고딕", size=9),
          align=Alignment(horizontal="left", vertical="center",
                          wrap_text=True, indent=1))
    return r + 2


# ---------- 섹션: 수업 계획 ----------
def render_schedule(ws: Worksheet, start_row: int, student: dict,
                    full_width: bool = False) -> int:
    """
    수업 계획 표 렌더링.
    full_width=False: A-E (평가 있음 → 우측에 달력)
    full_width=True:  A-M (평가 없음 → 폭 전체 사용)
    """
    prog = student.get("progress")
    if not prog:
        return start_row

    roles = prog["roles"]
    n_roles = len(roles)

    # 컬럼 매핑 결정
    if full_width:
        # 전체 폭: A=라벨, B-E = role1, F = spacer, G-I = role2 (2 roles 기준), 없으면 B-M = role1
        label_col = "A"
        if n_roles == 1:
            role_ranges = [("B", "M")]  # role1이 B-M 전체
        elif n_roles == 2:
            role_ranges = [("B", "G"), ("H", "M")]  # 2분할
        else:
            # 3 roles이상 — 균등 분할
            role_ranges = [("B", "E"), ("F", "I"), ("J", "M")][:n_roles]
    else:
        # 좁은 폭 (A-E):
        label_col = "A"
        if n_roles == 1:
            role_ranges = [("B", "E")]
        elif n_roles == 2:
            role_ranges = [("B", "C"), ("D", "E")]
        else:
            role_ranges = [("B", "C"), ("D", "E"), ("F", "F")][:n_roles]  # 특수 케이스

    r = start_row

    # 섹션 제목 행
    title_range = f"{label_col}{r}:{role_ranges[-1][1]}{r}"
    ws.row_dimensions[r].height = ROW_H_STANDARD
    merge(ws, title_range, f"{prog['month']}월 수업 계획",
          font=SECTION_TITLE_FONT, fill=NAVY_FILL)
    r += 1

    # 역할 라벨 행 (본행/선행)
    ws.row_dimensions[r].height = ROW_H_STANDARD
    set_cell(ws, f"{label_col}{r}", "", fill=LIGHT_FILL)
    for role, (cs, ce) in zip(roles, role_ranges):
        merge(ws, f"{cs}{r}:{ce}{r}", role["role"], font=ROLE_HEADER_FONT, fill=LIGHT_FILL)
    r += 1

    # 교재 행
    ws.row_dimensions[r].height = ROW_H_STANDARD
    set_cell(ws, f"{label_col}{r}", "교재", font=ROLE_HEADER_FONT, fill=LIGHT_FILL)
    for role, (cs, ce) in zip(roles, role_ranges):
        merge(ws, f"{cs}{r}:{ce}{r}", role.get("교재", ""), font=TEXTBOOK_FONT)
    r += 1

    # 주차별 (JSON 키 정규화)
    for role in roles:
        raw = role.get("주차별", {})
        role["주차별"] = {int(k): v for k, v in raw.items()}

    max_weeks = max((max(role["주차별"].keys()) if role["주차별"] else 0) for role in roles)

    for wk in range(1, max_weeks + 1):
        ws.row_dimensions[r].height = ROW_H_STANDARD
        set_cell(ws, f"{label_col}{r}", f"{wk}주차",
                 font=WEEK_LABEL_FONT, fill=LIGHT_FILL)
        for role, (cs, ce) in zip(roles, role_ranges):
            val = role["주차별"].get(wk, "")
            if isinstance(val, str):
                val = _normalize_month_format(val)
            merge(ws, f"{cs}{r}:{ce}{r}", val, font=CELL_FONT)
        r += 1

    return r + 1  # 여백 1행


# ---------- 섹션: 달력 ----------
def render_calendar(ws: Worksheet, start_row: int, year: int, month: int,
                    small: bool = True, start_col: str = "G",
                    end_col: str = "M") -> int:
    """
    small=True: 우측 작은 달력 (G-M)
    small=False: 전체 폭 큰 달력 (A-M)
    """
    cal = calendar.Calendar(firstweekday=6)
    weeks = cal.monthdayscalendar(year, month)

    sc_idx = ord(start_col) - ord("A") + 1
    ec_idx = ord(end_col) - ord("A") + 1
    total_cols = ec_idx - sc_idx + 1

    r = start_row
    # 제목
    ws.row_dimensions[r].height = ROW_H_STANDARD
    merge(ws, f"{start_col}{r}:{end_col}{r}", f"{year}년 {month}월",
          font=SECTION_TITLE_FONT, fill=NAVY_FILL)
    r += 1

    # 달력이 7요일 x 7일(Sun-Sat)이므로 대상 컬럼 수 7 필요
    # 만약 total_cols != 7: 요일을 비율로 배분 (A-M 13 cols → 2 cols/day 대략)
    day_labels = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

    if total_cols == 7:
        # 1 col per day
        day_ranges = [(get_column_letter(sc_idx + i), get_column_letter(sc_idx + i))
                      for i in range(7)]
    else:
        # 균등 분배
        base = total_cols // 7
        rem = total_cols % 7
        day_ranges = []
        cur = sc_idx
        for i in range(7):
            width = base + (1 if i < rem else 0)
            day_ranges.append((get_column_letter(cur), get_column_letter(cur + width - 1)))
            cur += width

    # 요일 헤더
    ws.row_dimensions[r].height = ROW_H_STANDARD
    for label, (cs, ce) in zip(day_labels, day_ranges):
        if cs == ce:
            set_cell(ws, f"{cs}{r}", label, font=CAL_HEADER_FONT, fill=NAVY_FILL)
        else:
            merge(ws, f"{cs}{r}:{ce}{r}", label, font=CAL_HEADER_FONT, fill=NAVY_FILL)
    r += 1

    # 날짜 셀. 큰 달력(small=False)은 bold 10pt, 작은 달력은 8pt
    row_height = ROW_H_STANDARD
    # 큰 달력용 폰트 생성
    if not small:
        day_font_default = Font(name="맑은 고딕", size=10, bold=True)
        day_font_sun = Font(name="맑은 고딕", size=10, bold=True, color="CC0000")
        day_font_sat = Font(name="맑은 고딕", size=10, bold=True, color="0066CC")
    else:
        day_font_default = CAL_DAY_FONT
        day_font_sun = CAL_SUN_FONT
        day_font_sat = CAL_SAT_FONT

    for week in weeks:
        ws.row_dimensions[r].height = row_height
        for di, day in enumerate(week):
            cs, ce = day_ranges[di]
            val = str(day) if day > 0 else ""
            if di == 0:
                font = day_font_sun
            elif di == 6:
                font = day_font_sat
            else:
                font = day_font_default
            if cs == ce:
                set_cell(ws, f"{cs}{r}", val, font=font, align=TOPLEFT)
            else:
                merge(ws, f"{cs}{r}:{ce}{r}", val, font=font, align=TOPLEFT)
        r += 1

    return r + 1


# ---------- 섹션: 평가 표 ----------
def _normalize_month_format(text: str) -> str:
    """'03월','04월' 등 앞 0 붙은 월 표기를 '3월','4월'로 정리."""
    return re.sub(r"\b0(\d)월", r"\1월", text)


def render_ai_comment(ws: Worksheet, start_row: int, comment: str) -> int:
    """AI 맞춤 코멘트 섹션."""
    if not comment:
        return start_row
    comment = _normalize_month_format(comment)
    r = start_row
    ws.row_dimensions[r].height = ROW_H_STANDARD
    merge(ws, f"A{r}:M{r}", "이번 달 선생님의 분석 및 코멘트",
          font=SECTION_TITLE_FONT, fill=NAVY_FILL)
    r += 1
    # 본문: 세로 가운데 정렬
    ws.row_dimensions[r].height = ROW_H_COMMENT_BODY
    merge(ws, f"A{r}:M{r}", comment,
          font=Font(name="맑은 고딕", size=9),
          align=Alignment(horizontal="left", vertical="center",
                          wrap_text=True, indent=1))
    return r + 2


def render_signature(ws: Worksheet, start_row: int) -> int:
    """학부모/담당 선생님 확인란."""
    r = start_row
    ws.row_dimensions[r].height = 30
    merge(ws, f"A{r}:F{r}", "학부모 확인: ___________________",
          font=Font(name="맑은 고딕", size=10),
          align=Alignment(horizontal="left", vertical="center", indent=1),
          border=False)
    merge(ws, f"G{r}:M{r}", "담당 선생님: ___________________",
          font=Font(name="맑은 고딕", size=10),
          align=Alignment(horizontal="left", vertical="center", indent=1),
          border=False)
    return r + 1


def render_eval_table(ws: Worksheet, start_row: int, title: str,
                      evals: list[dict], use_week: bool = True,
                      school_level: str = "", grade: str = "") -> int:
    if not evals:
        return start_row

    r = start_row
    ws.row_dimensions[r].height = ROW_H_STANDARD
    merge(ws, f"A{r}:M{r}", title, font=SECTION_TITLE_FONT, fill=NAVY_FILL)
    r += 1

    # 컬럼 헤더
    ws.row_dimensions[r].height = ROW_H_STANDARD
    set_cell(ws, f"A{r}", "", fill=LIGHT_FILL)
    merge(ws, f"B{r}:I{r}", "단원 명", font=ROLE_HEADER_FONT, fill=LIGHT_FILL)
    merge(ws, f"J{r}:K{r}", "점수", font=ROLE_HEADER_FONT, fill=LIGHT_FILL)
    merge(ws, f"L{r}:M{r}", "평균 정답률", font=ROLE_HEADER_FONT, fill=LIGHT_FILL)
    r += 1

    level_short = {"초등학교": "초", "중학교": "중", "고등학교": "고"}.get(school_level, "")

    for e in evals:
        ws.row_dimensions[r].height = ROW_H_STANDARD
        if use_week:
            week = e.get("주차")
            label = f"{week}주차" if week else ""
        else:
            label = f"{level_short}{grade}" if grade else ""

        set_cell(ws, f"A{r}", label, font=WEEK_LABEL_FONT, fill=LIGHT_FILL)

        unit = e["단원명"]
        retry_n = e.get("재시차수", 0)
        if retry_n > 0:
            retry_text = "재" * retry_n + "시"
            unit += f" ({retry_text})"
        unit = _normalize_month_format(unit)
        merge(ws, f"B{r}:I{r}", unit, font=CELL_FONT, align=LEFT)

        score = e["점수"]
        score_font = RED_FONT if (score is not None and score < 70) else CELL_FONT
        merge(ws, f"J{r}:K{r}", score, font=score_font)

        avg = e.get("전국평균", 75)
        merge(ws, f"L{r}:M{r}", f"{avg}%", font=CELL_FONT)
        r += 1

    return r + 1


def load_notice(year: int, month: int) -> str:
    """월별 공지사항 파일 로드. 없으면 기본 placeholder."""
    base = Path(__file__).parent.parent
    notice_path = base / "output" / f"notice_{year}_{month:02d}.txt"
    if notice_path.exists():
        return notice_path.read_text(encoding="utf-8").strip()
    # 파일 없으면 placeholder
    return (f"● \n"
            f"● \n"
            f"● \n\n"
            f"(공지사항은 output/notice_{year}_{month:02d}.txt 파일에 입력하세요)")


# ---------- 학생 시트 ----------
def render_student_sheet(wb: Workbook, student: dict, year: int, month: int,
                         notice_text: str = ""):
    sheet_name = f"{student['code']}_{student['name']}"[:31]
    ws = wb.create_sheet(sheet_name)
    setup_page(ws)

    # 레이아웃 B (평가 없음, 큰 달력): 원장님 실측 폭으로 오버라이드
    if not student["has_evaluations"]:
        for col in "ABCDEFGHIJKL":
            ws.column_dimensions[col].width = 6.08  # 80px
        ws.column_dimensions["M"].width = 12.75  # 140px

    # 타이틀 + NAME/DATE
    next_row = render_title(ws)
    next_row = render_name_date(ws, next_row, student, year, month)
    next_row = render_warnings(ws, next_row, student.get("warnings", []))

    # 공지사항 (학생 이름 아래, 진도계획표 위)
    next_row = render_notice(ws, next_row, notice_text)

    has_evals = student["has_evaluations"]
    has_progress = student["has_progress"]

    if has_evals:
        # 레이아웃 A: 좌측 수업계획 + 우측 달력 (같은 시작 row)
        schedule_start = next_row
        cal_start = next_row

        sched_end = schedule_start
        if has_progress:
            sched_end = render_schedule(ws, schedule_start, student, full_width=False)
        cal_end = render_calendar(ws, cal_start, year, month, small=True,
                                   start_col="G", end_col="M")

        next_row = max(sched_end, cal_end) + 1

        # 평가 표
        if student["weekly_evals"]:
            next_row = render_eval_table(ws, next_row, "주간 평가",
                                          student["weekly_evals"], use_week=True)
        if student["unit_evals"]:
            next_row = render_eval_table(ws, next_row, "선행 평가",
                                          student["unit_evals"], use_week=False,
                                          school_level=student["school_level"],
                                          grade=student["grade"])
    else:
        # 레이아웃 B: 전체 폭 수업계획 + 큰 달력
        if has_progress:
            next_row = render_schedule(ws, next_row, student, full_width=True)
        next_row = render_calendar(ws, next_row, year, month, small=False,
                                    start_col="A", end_col="M")

    # AI 코멘트 (있으면)
    if student.get("ai_comment"):
        next_row = render_ai_comment(ws, next_row, student["ai_comment"])


# ---------- 메인 ----------
def main():
    if len(sys.argv) < 3:
        print("사용법: python render_gatongmun.py <연도> <월>")
        sys.exit(1)
    year = int(sys.argv[1])
    month = int(sys.argv[2])

    base = Path(__file__).parent.parent
    merged_json = base / "output" / f"merged_{year}_{month:02d}.json"
    if not merged_json.exists():
        print(f"❌ {merged_json.name} 없음")
        sys.exit(1)

    with merged_json.open(encoding="utf-8") as f:
        data = json.load(f)

    notice_text = load_notice(year, month)
    wb = Workbook()
    wb.remove(wb.active)

    for student in data["students"]:
        render_student_sheet(wb, student, year, month, notice_text=notice_text)

    out_path = base / "output" / f"가통문_{year}_{month:02d}.xlsx"
    try:
        wb.save(out_path)
    except PermissionError:
        # 파일이 열려있으면 v2, v3 ... 생성
        for i in range(2, 100):
            alt = base / "output" / f"가통문_{year}_{month:02d}_v{i}.xlsx"
            try:
                wb.save(alt)
                out_path = alt
                print(f"   ⚠️ 기존 파일이 열려있어 새 파일로 저장: {alt.name}")
                break
            except PermissionError:
                continue

    # 레이아웃 분포
    n_a = sum(1 for s in data["students"] if s["has_evaluations"])
    n_b = len(data["students"]) - n_a

    print(f"✅ 가통문 생성: {len(data['students'])}명 ({len(wb.sheetnames)}시트)")
    print(f"   레이아웃 A (평가 있음): {n_a}명")
    print(f"   레이아웃 B (평가 없음): {n_b}명")
    print(f"   파일: {out_path}")


if __name__ == "__main__":
    main()
