"""
진도계획표 Google Sheet 파서 (단순화 버전)

핵심 규칙:
- 각 학생 블록의 **위쪽 블록(첫 월 블록, 약 row+3 ~ row+9)만** 가통문용 데이터
- 아래 행들은 선생님 참고용 → 무시
- 블록 헤더의 연도는 무시 → 사용자가 지정한 연도+월 사용
- 4주차/5주차는 데이터 유무로 자동 판정
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

ROLE_LABELS = {"본행", "선행", "본행1", "본행2", "선행1", "선행2"}
WEEK_RE = re.compile(r"^(\d+)\s*주차$")
MONTH_RE = re.compile(r"^(\d+)\s*월$")


def norm(val: Any) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).strip())


def is_student_name(val: Any) -> bool:
    s = norm(val)
    return bool(re.match(r"^(초\d|중\d|고\d)[A-Za-z]*\d*\s+[가-힣]{2,4}$", s))


def parse_student_code(val: Any) -> tuple[str, str] | None:
    s = norm(val)
    m = re.match(r"^((?:초|중|고)\d[A-Za-z]*\d*)\s+([가-힣]{2,4})$", s)
    return (m.group(1), m.group(2)) if m else None


def parse_tab(ws, target_month: int | None = None) -> list[dict]:
    """
    탭 하나에서 학생 블록 스캔 → 각 학생의 **첫 번째 월 블록**만 추출.
    target_month가 지정되면 해당 월과 매칭되는 블록만 반환.
    """
    max_row, max_col = ws.max_row, ws.max_column

    # 전체 셀 → 2D 리스트 (1-based)
    grid = [[None] * (max_col + 1)]
    for r in range(1, max_row + 1):
        row = [None]
        for c in range(1, max_col + 1):
            row.append(ws.cell(r, c).value)
        grid.append(row)

    # 1단계: 학생 블록 헤더 찾기 (학생명 셀 위치)
    blocks = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if is_student_name(grid[r][c]):
                parsed = parse_student_code(grid[r][c])
                if parsed:
                    blocks.append({"row": r, "col": c, "code": parsed[0], "name": parsed[1]})

    # 같은 이름이 여러 번 나오면 (과거 학년 블록) 맨 위(첫 등장)만 유지
    seen = {}
    unique_blocks = []
    for b in blocks:
        key = b["name"]
        if key not in seen:
            seen[key] = True
            unique_blocks.append(b)
    blocks = unique_blocks

    # 2단계: 각 블록에서 위쪽 월 블록 데이터만 추출
    results = []
    for b in blocks:
        anchor = b["col"]  # 학생명 셀 컬럼

        # 첫 번째 월 라벨 위치 찾기 (anchor 주변 넓은 범위에서)
        month_row, month_val, month_col = None, None, None
        for dr in range(1, 8):
            r = b["row"] + dr
            if r > max_row:
                break
            for cc in range(max(1, anchor - 2), min(max_col + 1, anchor + 5)):
                m = MONTH_RE.match(norm(grid[r][cc]))
                if m:
                    month_row = r
                    month_val = int(m.group(1))
                    month_col = cc
                    break
            if month_row:
                break

        if month_row is None:
            continue

        if target_month is not None and month_val != target_month:
            continue

        # 월 라벨 위치 기준 상대 오프셋:
        #   month_col+1 = sub-label 컬럼 (교재/담당T/N주차)
        #   month_col+2 = 첫 번째 역할 값 컬럼 (본행)
        #   month_col+4 = 두 번째 역할 값 컬럼 (선행 또는 본행2)
        sub_col = month_col + 1

        # Role 컬럼 찾기: 블록 헤더 +1~+3 행에서 본행/선행 라벨 탐색
        role_cols: dict[int, str] = {}
        for dr in range(1, 4):
            r = b["row"] + dr
            if r > max_row or r >= month_row:
                break
            for cc in range(max(1, month_col), min(max_col + 1, month_col + 7)):
                lbl = norm(grid[r][cc])
                if lbl in ROLE_LABELS:
                    role_cols[cc] = lbl
            if role_cols:
                break

        # Role label이 없는 케이스 (본행만 있는 학생 등) → 기본값 생성
        if not role_cols:
            # 교재/주차별 row에서 값 있는 컬럼을 탐색
            candidate_cols = []
            for cc in (month_col + 2, month_col + 4):
                if cc > max_col:
                    continue
                # 교재 행 또는 주차 행 중 하나라도 값이 있으면 포함
                has_data = False
                for rr in range(month_row, min(month_row + 8, max_row + 1)):
                    if norm(grid[rr][cc]):
                        has_data = True
                        break
                if has_data:
                    candidate_cols.append(cc)
            if candidate_cols:
                role_cols = {candidate_cols[0]: "본행"}
                if len(candidate_cols) >= 2:
                    role_cols[candidate_cols[1]] = "선행"

        if not role_cols:
            continue

        role_data = {c: {"role": role, "교재": "", "주차별": {}} for c, role in role_cols.items()}

        for rr in range(month_row, min(month_row + 8, max_row + 1)):
            label = norm(grid[rr][sub_col]) if sub_col <= max_col else ""

            # 다음 월 라벨이 나오면 중단
            if rr != month_row:
                next_month_found = False
                for cc in range(max(1, month_col - 2), min(max_col + 1, month_col + 5)):
                    if MONTH_RE.match(norm(grid[rr][cc])):
                        next_month_found = True
                        break
                if next_month_found:
                    break

            if label == "교재":
                for c in role_data:
                    role_data[c]["교재"] = norm(grid[rr][c])
            elif label == "담당T":
                pass
            else:
                wm = WEEK_RE.match(label)
                if wm:
                    week_n = int(wm.group(1))
                    for c in role_data:
                        val = norm(grid[rr][c])
                        if val:
                            role_data[c]["주차별"][week_n] = val

        roles = []
        for c, rd in role_data.items():
            if rd["교재"] or rd["주차별"]:
                roles.append({
                    "role": rd["role"],
                    "교재": rd["교재"],
                    "주차별": rd["주차별"],
                })

        if roles:
            results.append({
                "code": b["code"],
                "name": b["name"],
                "month": month_val,
                "roles": roles,
            })

    return results


def parse_progress_workbook(
    xlsx_path: str | Path,
    target_year: int,
    target_month: int,
) -> dict:
    """모든 탭 파싱 → target_month 데이터만 추출. 연도는 사용자 지정."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_students = []

    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        tab_results = parse_tab(ws, target_month)
        for s in tab_results:
            s["tab"] = tab_name
            s["year"] = target_year
        all_students.extend(tab_results)

    # 같은 이름 중복 제거 (첫 등장 = 최신 탭 기준)
    seen = {}
    deduped = []
    for s in all_students:
        key = s["name"]
        if key not in seen:
            seen[key] = True
            deduped.append(s)

    return {
        "source_file": Path(xlsx_path).name,
        "year": target_year,
        "month": target_month,
        "students": deduped,
    }


def main():
    if len(sys.argv) < 4:
        print("사용법: python parse_progress.py <xlsx경로> <연도> <월>")
        print("예시:   python parse_progress.py 진도계획표.xlsx 2026 4")
        sys.exit(1)

    xlsx = Path(sys.argv[1])
    year = int(sys.argv[2])
    month = int(sys.argv[3])

    data = parse_progress_workbook(xlsx, year, month)

    out_path = xlsx.parent / f"progress_{year}_{month:02d}.json"
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✅ 파싱 완료: {xlsx.name}  → {year}년 {month}월")
    print(f"   학생 수: {len(data['students'])}")
    for s in data["students"]:
        roles_desc = []
        for r in s["roles"]:
            weeks = sorted(r["주차별"].keys())
            week_range = f"{min(weeks)}~{max(weeks)}주차" if weeks else "주차없음"
            roles_desc.append(f"{r['role']}({r['교재'][:12] or '교재미정'}, {week_range})")
        print(f"   {s['code']:8s} {s['name']:5s} | {' / '.join(roles_desc)}")
    print(f"   출력: {out_path}")


if __name__ == "__main__":
    main()
